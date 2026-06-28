"""
build_leads_excel.py — Scorpion Global Solutions Trigger-Lead Excel Generator
=============================================================================

Usage (from another script or inline Python):

    from build_leads_excel import build_excel

    leads = [
        (
            "Company Name",           # 0  Company
            "Jane Doe",               # 1  Decision-Maker
            "CEO",                    # 2  Title
            "Multifamily (value-add)",# 3  Industry / Asset
            "Dallas, TX",             # 4  Location
            "Bridge / Commercial RE", # 5  Loan Product Fit
            "Acquired 200-unit complex for $45M on Jun 15 2026; value-add capex planned.",  # 6  Trigger Event
            "$45M",                   # 7  Est. Deal Size
            78,                       # 8  Lead Score  (int 0–100)
            "A",                      # 9  Grade       ("A", "B", or "C")
            None,                     # 10 Verified Email
            None,                     # 11 Email Status
            "jdoe@company.com",       # 12 Best-Effort Email (unverified)
            "https://linkedin.com/in/jane-doe",  # 13 LinkedIn
            None,                     # 14 Mobile Phone
            None,                     # 15 Other Public Contact
            "https://therealdeal.com/...",       # 16 Source URL
        ),
        # ... more leads
    ]

    build_excel(leads, "/path/to/output.xlsx", batch_label="June 2026 Batch")

The function returns the output path on success and raises on error.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Colour palette ──────────────────────────────────────────────────────────
DARK_BG   = "1A1A2E"
MID_BG    = "16213E"
ACCENT    = "0F3460"
GOLD      = "E94560"
WHITE     = "FFFFFF"
LIGHT_ROW = "F0F4FF"
ALT_ROW   = "E8EDF8"
GRADE_A   = "1E8449"
GRADE_B   = "D4AC0D"
GRADE_C   = "CA6F1E"

GRADE_COLORS = {"A": GRADE_A, "B": GRADE_B, "C": GRADE_C}

COL_WIDTHS = [32, 26, 30, 26, 22, 26, 60, 16, 8, 8, 34, 16, 34, 40, 20, 30, 55]

HEADERS = [
    "Company", "Decision-Maker", "Title", "Industry / Asset", "Location",
    "Loan Product Fit", "Trigger Event", "Est. Deal Size", "Lead Score",
    "Grade", "Verified Email", "Email Status", "Best-Effort Email (unverified)",
    "LinkedIn", "Mobile Phone", "Other Public Contact", "Source",
]


def _thin_border():
    return Border(
        right=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )


def build_excel(leads: list, output_path: str, batch_label: str = "New Leads Batch") -> str:
    """
    Build a two-tab Excel workbook from a list of lead tuples.

    Parameters
    ----------
    leads       : list of 17-element tuples (see module docstring for schema)
    output_path : absolute path for the output .xlsx file
    batch_label : short label shown in the title banner and summary tab

    Returns
    -------
    str : output_path on success
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1 — title banner
    ws.merge_cells("A1:Q1")
    ws["A1"] = f"SCORPION GLOBAL SOLUTIONS  –  Trigger-Based Commercial-Finance Loan Leads  |  {batch_label}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Row 2 — sub-header
    ws.merge_cells("A2:Q2")
    ws["A2"] = (
        f"{len(leads)} leads · trigger events from last 2 weeks · "
        "no duplicates from master list · scored A/B/C"
    )
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="CCCCCC")
    ws["A2"].fill = PatternFill("solid", fgColor=MID_BG)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Row 3 — column headers
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="medium", color=GOLD),
            right=Side(style="thin", color="334477"),
        )
    ws.row_dimensions[3].height = 30

    # Data rows
    for i, lead in enumerate(leads):
        row = i + 4
        fill = PatternFill("solid", fgColor=LIGHT_ROW if i % 2 == 0 else ALT_ROW)

        for col, val in enumerate(lead, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _thin_border()

        # Grade badge
        grade = lead[9]
        gc = ws.cell(row=row, column=10)
        gc.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        gc.fill = PatternFill("solid", fgColor=GRADE_COLORS.get(grade, GRADE_C))
        gc.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 60

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:Q{3 + len(leads)}"

    # ── Summary tab ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20

    title_cell = ws2.cell(row=1, column=1, value=f"BATCH SUMMARY – {batch_label}")
    title_cell.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=DARK_BG)
    ws2.merge_cells("A1:B1")
    ws2.row_dimensions[1].height = 28

    grade_a = sum(1 for l in leads if l[9] == "A")
    grade_b = sum(1 for l in leads if l[9] == "B")
    grade_c = sum(1 for l in leads if l[9] == "C")

    product_counts: dict[str, int] = {}
    for l in leads:
        product_counts[l[5]] = product_counts.get(l[5], 0) + 1

    summary_rows = [
        ("Total Leads", len(leads)),
        ("Grade A Leads", grade_a),
        ("Grade B Leads", grade_b),
        ("Grade C Leads", grade_c),
        ("", ""),
        ("LOAN PRODUCT BREAKDOWN", ""),
    ] + sorted(product_counts.items(), key=lambda x: -x[1])

    for r, (label, value) in enumerate(summary_rows, 2):
        ca = ws2.cell(row=r, column=1, value=label)
        cb = ws2.cell(row=r, column=2, value=value)
        if label in ("LOAN PRODUCT BREAKDOWN",):
            ca.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            ca.fill = PatternFill("solid", fgColor=ACCENT)
            cb.fill = PatternFill("solid", fgColor=ACCENT)
        else:
            ca.font = Font(name="Calibri", size=10)
            cb.font = Font(name="Calibri", bold=True, size=10)
        ws2.row_dimensions[r].height = 18

    wb.save(output_path)
    return output_path


# ── CLI convenience ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("build_leads_excel.py — import build_excel() from this module to use it.")
    print("See module docstring for the lead tuple schema.")
